import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


def to_int_or_none(value):
    try:
        value = int(value)
        if value > 0:
            return value
    except (TypeError, ValueError):
        return None


def is_header_row(row):
    if not row:
        return False 
    for cell in row[:3]:
        if cell is not None and isinstance(cell, str) and cell.strip():
            return True
    return False


def get_graph_from_excel(file_obj, filename: str) -> dict:
    workbook = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    edges = []
    nodes = set()  # повторяющиеся узлы будут игнорированы

    header_rows = 0
    for row in sheet.iter_rows(max_row=3, values_only=True):
        if is_header_row(row):
            header_rows += 1
        else:
            break
    start_row = header_rows + 1

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        source, target, node = map(to_int_or_none, row)

        if node is not None:
            nodes.add(node)
        if source is not None and target is not None:
            edges.append({"source": source, "target": target})

    filtered_edges = [e for e in edges if e["source"] in nodes and e["target"] in nodes]

    return {
        "id": str(uuid.uuid4()),
        "nodes": list(nodes),
        "edges": filtered_edges,
        "upload_time": datetime.now(timezone.utc).isoformat(),
        "filename": Path(filename).stem or "<Без названия>",
    }
